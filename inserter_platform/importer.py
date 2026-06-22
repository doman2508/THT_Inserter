from __future__ import annotations

import re
import unicodedata
from collections import OrderedDict
from dataclasses import dataclass
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

from . import db


@dataclass(frozen=True)
class BomItem:
    designator: str
    value: str
    medcom_index: str
    source_row: int
    note: str = ""


@dataclass(frozen=True)
class PlacementPoint:
    designator: str
    x: float
    y: float
    rotation: float | None
    source_row: int


def normalize_header(value: object) -> str:
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    return re.sub(r"[^a-z0-9]+", "", text.lower())


def cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).replace("\xa0", " ").strip()


def parse_float(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = cell_text(value).replace(",", ".")
    match = re.search(r"[-+]?\d+(?:\.\d+)?", text)
    if not match:
        return None
    return float(match.group(0))


# Standard designators are simple (R62, U3, J1), but some customer BOMs use
# hierarchical names such as RP.A0.1. Keep the token strict enough to avoid notes.
DESIGNATOR_TOKEN_RE = re.compile(r"^(?=[A-Z][A-Z0-9.']*$)(?=.*\d)[A-Z][A-Z0-9']*(?:\.[A-Z0-9']+)*$")
PLACEMENT_TOKEN_RE = re.compile(r"^[A-Z][A-Z0-9']*(?:\.[A-Z0-9']+)*$")


def clean_designator_token(value: str) -> str:
    value = str(value).translate(str.maketrans({
        "·": ".",
        "•": ".",
        "․": ".",
        "‧": ".",
        "∙": ".",
        "．": ".",
    }))
    return value.strip().strip("()[]{}.,:").replace("’", "'").upper()


def parse_designator_cell(value: object) -> tuple[list[str], str]:
    raw_text = cell_text(value)
    if not raw_text:
        return [], ""

    designators: list[str] = []
    note_parts: list[str] = []
    for chunk in re.split(r"[,;]+", raw_text):
        tokens = [token for token in re.split(r"[\s+/]+", chunk.strip()) if token]
        if not tokens:
            continue

        chunk_designators: list[str] = []
        note_start_index: int | None = None
        for index, token in enumerate(tokens):
            clean_token = clean_designator_token(token)
            if DESIGNATOR_TOKEN_RE.match(clean_token):
                chunk_designators.append(clean_token)
                continue
            note_start_index = index
            break

        if chunk_designators:
            designators.extend(chunk_designators)
            if note_start_index is not None:
                note_parts.append(" ".join(tokens[note_start_index:]).strip())

    return designators, " | ".join(part for part in note_parts if part)


def split_designators(value: object) -> list[str]:
    parsed_designators, _ = parse_designator_cell(value)
    return parsed_designators


def split_placement_designators(value: object) -> list[str]:
    raw_text = cell_text(value)
    if not raw_text:
        return []
    designators: list[str] = []
    for token in re.split(r"[,;\s+/]+", raw_text):
        clean_token = clean_designator_token(token)
        if clean_token and PLACEMENT_TOKEN_RE.match(clean_token):
            designators.append(clean_token)
    return designators


def natural_designator_key(value: str) -> tuple[Any, ...]:
    parts = re.split(r"(\d+)", value.upper())
    return tuple(int(part) if part.isdigit() else part for part in parts if part != "")


def read_xlsx_table(file_bytes: bytes, required_headers: dict[str, str]) -> list[dict[str, Any]]:
    workbook = load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
    try:
        worksheet = workbook[workbook.sheetnames[0]]
        rows = list(worksheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    required_normalized = {normalize_header(header): field for header, field in required_headers.items()}
    header_row_index = None
    column_map: dict[int, str] = {}

    for index, row in enumerate(rows[:50]):
        normalized_headers = [normalize_header(cell) for cell in row]
        if all(header in normalized_headers for header in required_normalized):
            header_row_index = index
            for column_index, normalized in enumerate(normalized_headers):
                if normalized in required_normalized:
                    column_map[column_index] = required_normalized[normalized]
            break

    if header_row_index is None:
        expected = ", ".join(required_headers.keys())
        raise ValueError(f"Nie znaleziono wymaganych kolumn: {expected}")

    table_rows: list[dict[str, Any]] = []
    for row_number, row in enumerate(rows[header_row_index + 1 :], start=header_row_index + 2):
        item = {field: row[column_index] if column_index < len(row) else None for column_index, field in column_map.items()}
        if any(cell_text(value) for value in item.values()):
            item["source_row"] = row_number
            table_rows.append(item)
    return table_rows


def parse_bom(file_bytes: bytes) -> tuple[list[BomItem], dict[str, Any]]:
    rows = read_xlsx_table(
        file_bytes,
        {
            "Desygnator": "designators",
            "Wartość": "value",
            "Indeks Medcom": "medcom_index",
        },
    )
    items: list[BomItem] = []
    empty_designator_rows: list[int] = []

    for row in rows:
        designators, designator_note = parse_designator_cell(row["designators"])
        if not designators:
            empty_designator_rows.append(int(row["source_row"]))
            continue
        for designator in designators:
            items.append(
                BomItem(
                    designator=designator,
                    value=cell_text(row["value"]),
                    medcom_index=cell_text(row["medcom_index"]),
                    source_row=int(row["source_row"]),
                    note=designator_note,
                )
            )

    return items, {
        "bomRows": len(rows),
        "bomDesignators": len(items),
        "bomRowsWithoutDesignators": empty_designator_rows,
    }


def parse_placement(file_bytes: bytes) -> tuple[list[PlacementPoint], dict[str, Any]]:
    rows = read_xlsx_table(
        file_bytes,
        {
            "Desygnator": "designator",
            "X": "x",
            "Y": "y",
            "Rotacja": "rotation",
        },
    )
    points: list[PlacementPoint] = []
    invalid_rows: list[int] = []

    for row in rows:
        designators = split_placement_designators(row["designator"])
        x = parse_float(row["x"])
        y = parse_float(row["y"])
        rotation = parse_float(row["rotation"])
        if len(designators) != 1 or x is None or y is None:
            invalid_rows.append(int(row["source_row"]))
            continue
        points.append(
            PlacementPoint(
                designator=designators[0],
                x=x,
                y=y,
                rotation=rotation,
                source_row=int(row["source_row"]),
            )
        )

    return points, {
        "ppRows": len(rows),
        "ppPoints": len(points),
        "ppInvalidRows": invalid_rows,
    }


def points_from_placement(pp_points: list[PlacementPoint]) -> tuple[list[dict[str, Any]], list[str]]:
    pp_by_designator: OrderedDict[str, PlacementPoint] = OrderedDict()
    pp_duplicates: list[str] = []
    for point in pp_points:
        if point.designator in pp_by_designator:
            pp_duplicates.append(point.designator)
            continue
        pp_by_designator[point.designator] = point
    return [
        {
            "designator": point.designator,
            "x": point.x,
            "y": point.y,
            "rotation": point.rotation,
        }
        for point in pp_by_designator.values()
    ], pp_duplicates


def prepare_point_supplement(*, pp_bytes: bytes) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    pp_points, pp_summary = parse_placement(pp_bytes)
    points, duplicates = points_from_placement(pp_points)
    return points, {
        "sourceType": "pp-point-supplement",
        **pp_summary,
        "uniquePpPoints": len(points),
        "ppDuplicateDesignators": sorted(set(duplicates), key=natural_designator_key),
    }


def build_steps_and_points(bom_items: list[BomItem], pp_points: list[PlacementPoint]) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, Any]]:
    bom_by_designator: OrderedDict[str, BomItem] = OrderedDict()
    bom_duplicates: list[str] = []
    bom_conflicts: list[str] = []

    for item in bom_items:
        previous = bom_by_designator.get(item.designator)
        if previous:
            bom_duplicates.append(item.designator)
            if (previous.value, previous.medcom_index) != (item.value, item.medcom_index):
                bom_conflicts.append(item.designator)
            continue
        bom_by_designator[item.designator] = item

    pp_points_dicts, pp_duplicates = points_from_placement(pp_points)
    pp_by_designator = {
        point["designator"]: PlacementPoint(
            designator=point["designator"],
            x=float(point["x"]),
            y=float(point["y"]),
            rotation=point.get("rotation"),
            source_row=0,
        )
        for point in pp_points_dicts
    }

    missing_in_pp = [designator for designator in bom_by_designator if designator not in pp_by_designator]
    extra_in_pp = [designator for designator in pp_by_designator if designator not in bom_by_designator]
    matched_designators = [designator for designator in bom_by_designator if designator in pp_by_designator]

    groups: OrderedDict[tuple[str, str, bool], list[str]] = OrderedDict()
    for designator in bom_by_designator:
        item = bom_by_designator[designator]
        has_point = designator in pp_by_designator
        groups.setdefault((item.value, item.medcom_index, has_point), []).append(designator)

    steps: list[dict[str, Any]] = []
    for step_no, ((value, medcom_index, has_point), designators) in enumerate(groups.items(), start=1):
        sorted_designators = sorted(designators, key=natural_designator_key)
        missing_group_designators = [] if has_point else sorted_designators
        note_parts = [
            bom_by_designator[designator].note
            for designator in sorted_designators
            if bom_by_designator[designator].note
        ]
        if missing_group_designators:
            note_parts.append(f"Brak punktu P&P: {', '.join(missing_group_designators)}")
        notes = " | ".join(dict.fromkeys(note_parts))
        steps.append(
            {
                "step_no": step_no,
                "designators": ",".join(sorted_designators),
                "value": value,
                "medcom_index": medcom_index,
                "quantity": len(sorted_designators),
                "seconds": None,
                "notes": notes,
            }
        )

    points = [
        {
            "designator": designator,
            "x": pp_by_designator[designator].x,
            "y": pp_by_designator[designator].y,
            "rotation": pp_by_designator[designator].rotation,
        }
        for designator in matched_designators
    ]

    return steps, points, {
        "matchedDesignators": len(matched_designators),
        "createdSteps": len(steps),
        "createdPoints": len(points),
        "missingInPp": missing_in_pp,
        "missingInPpImportedAsSteps": True,
        "extraInPp": extra_in_pp,
        "bomDuplicateDesignators": sorted(set(bom_duplicates), key=natural_designator_key),
        "bomConflictingDesignators": sorted(set(bom_conflicts), key=natural_designator_key),
        "ppDuplicateDesignators": sorted(set(pp_duplicates), key=natural_designator_key),
    }


def prepare_prepared_import(*, bom_bytes: bytes, pp_bytes: bytes) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, Any]]:
    bom_items, bom_summary = parse_bom(bom_bytes)
    pp_points, pp_summary = parse_placement(pp_bytes)
    steps, points, merge_summary = build_steps_and_points(bom_items, pp_points)
    if not steps:
        raise ValueError("Import nie utworzył żadnego kroku. Sprawdź zgodność desygnatorów BOM i P&P.")

    summary = {
        "sourceType": "prepared-xlsx",
        **bom_summary,
        **pp_summary,
        **merge_summary,
    }
    return steps, points, summary


def import_prepared_project(
    *,
    name: str,
    board_width: float | None,
    board_height: float | None,
    bom_bytes: bytes,
    pp_bytes: bytes,
) -> tuple[dict[str, Any], dict[str, Any]]:
    steps, points, summary = prepare_prepared_import(bom_bytes=bom_bytes, pp_bytes=pp_bytes)
    project = db.create_project_from_import(
        name=name,
        board_width=board_width,
        board_height=board_height,
        steps=steps,
        points=points,
        summary=summary,
    )
    return project, summary
